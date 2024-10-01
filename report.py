import os
import requests
import pandas as pd
from datetime import datetime, timedelta
from azure.storage.blob import BlobServiceClient, BlobSasPermissions, generate_blob_sas
from azure.core.exceptions import ResourceExistsError
from azure.identity import DefaultAzureCredential
import logging
import argparse
from io import BytesIO
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

# Import required classes from azure.cosmos
from azure.cosmos import CosmosClient, PartitionKey, exceptions

from db import get_mongo_client, get_cosmos_client, append_new_data, save_data

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def sanitize_name(name):
    """Replaces non-alphanumeric characters with dashes and converts to lowercase."""
    return ''.join(char if char.isalnum() else '-' for char in name).lower()

def retrieve_and_process_stats(owner, repo, filename,
                               db_connection_string, db_type,
                               azure_storage_connection_string,
                               output_format, token, use_managed_identity):
    try:
        base_url = f"https://api.github.com/repos/{owner}/{repo}"
        sanitized_repo = sanitize_name(repo)
        filename = filename or f"{owner}-{sanitized_repo}"
        base_filename = os.path.splitext(filename)[0]

        logger.info("Starting data retrieval and processing")
        logger.info(f"Repository: {owner}/{repo}")
        logger.info(f"Filename: {filename}")
        logger.info(f"Database type: {db_type}")

        # Create database client
        db_client = get_db_client(db_connection_string, db_type)

        # Fetch repository info
        repo_info = get_github_repo_info(base_url, token)

        if repo_info:
            about_data = process_about_data(repo_info, owner, repo)
            about_df = pd.DataFrame([about_data])
        else:
            about_df = pd.DataFrame(columns=['Repo Owner', 'Repo Name', 'About'])

        # Fetch and process data from GitHub API
        views_data = get_github_data(f"{base_url}/traffic/views", token)
        clones_data = get_github_data(f"{base_url}/traffic/clones", token)
        referrers_data = get_github_data(f"{base_url}/traffic/popular/referrers", token)
        popular_content_data = get_github_data(f"{base_url}/traffic/popular/paths", token)
        stars_data = get_stars_data(base_url, token)
        forks_data = get_forks_data(base_url, token)

        logger.info(f"Fetched data for repository: {owner}/{repo}")

        # Convert the processed data to DataFrames
        referrers_df = pd.DataFrame(process_referrers_data(referrers_data, owner, repo))
        popular_content_df = pd.DataFrame(process_popular_content_data(popular_content_data, owner, repo))
        stars_df = pd.DataFrame(process_stars_data(stars_data, owner, repo, db_client, db_type))
        forks_df = pd.DataFrame(process_forks_data(forks_data, owner, repo, db_client, db_type))

        # Process and append the new data
        traffic_data_processed = process_traffic_data(views_data, owner, repo)
        traffic_df = append_new_data(db_client, sanitized_repo, "TrafficStats", traffic_data_processed, 'Date', db_type)
        clones_data_processed = process_clones_data(clones_data, owner, repo)
        clones_df = append_new_data(db_client, sanitized_repo, "GitClones", clones_data_processed, 'Date', db_type)

        # Save to database
        ensure_collection_exists(db_client, sanitized_repo, "About", db_type)
        save_data(db_client, sanitized_repo, "About", about_df.to_dict('records'), db_type, owner, repo)

        ensure_collection_exists(db_client, sanitized_repo, "TrafficStats", db_type)
        save_data(db_client, sanitized_repo, "TrafficStats", traffic_df.to_dict('records'), db_type, owner, repo)

        ensure_collection_exists(db_client, sanitized_repo, "GitClones", db_type)
        save_data(db_client, sanitized_repo, "GitClones", clones_df.to_dict('records'), db_type, owner, repo)

        ensure_collection_exists(db_client, sanitized_repo, "Stars", db_type)
        save_data(db_client, sanitized_repo, "Stars", stars_df.to_dict('records'), db_type, owner, repo)

        ensure_collection_exists(db_client, sanitized_repo, "Forks", db_type)
        save_data(db_client, sanitized_repo, "Forks", forks_df.to_dict('records'), db_type, owner, repo)

        logger.info("Data saved to database")

        # Define the dataframes dictionary
        dataframes = {
            'About': about_df,
            'TrafficStats': traffic_df,
            'GitClones': clones_df,
            'Stars': stars_df,
            'Forks': forks_df
        }

        # Handle Azure Blob Storage uploads
        handle_azure_blob_storage(azure_storage_connection_string, dataframes, base_filename, output_format, use_managed_identity)

    except Exception as e:
        logger.exception(f"An error occurred: {e}")

def ensure_collection_exists(db_client, repo, collection_name, db_type):
    """Ensure that the collection exists in the database."""
    if db_type == "mongodb":
        db = db_client[repo]
        if collection_name not in db.list_collection_names():
            db.create_collection(collection_name)
    elif db_type == "cosmosdb":
        db = db_client.get_database_client(repo)
        try:
            db.create_container_if_not_exists(id=collection_name, partition_key=PartitionKey(path='/id'))
        except Exception as e:
            logger.error(f"Error creating collection '{collection_name}': {e}")

def get_db_client(connection_string, db_type):
    """Get the appropriate database client based on the database type."""
    if db_type == "mongodb":
        return get_mongo_client(connection_string)
    elif db_type == "cosmosdb":
        return get_cosmos_client(connection_string)
    else:
        raise ValueError(f"Unsupported database type: {db_type}")

def handle_azure_blob_storage(connection_string, dataframes, base_filename, output_format, use_managed_identity):
    """Handle uploading dataframes to Azure Blob Storage."""
    if connection_string:
        container_name = sanitize_name(base_filename)

        # Upload JSON files directly to Azure Blob Storage
        if output_format in ['json', 'all']:
            upload_json_to_azure_blob(connection_string, dataframes, base_filename, container_name, use_managed_identity)

        # Upload Excel file directly to Azure Blob Storage
        if output_format in ['excel', 'all']:
            upload_excel_to_azure_blob(connection_string, dataframes, base_filename, container_name, use_managed_identity)
    else:
        save_files_locally(dataframes, base_filename, output_format)

def upload_json_to_azure_blob(connection_string, dataframes, base_filename, container_name, use_managed_identity):
    """Upload JSON dataframes to Azure Blob Storage."""
    for df_name, df in dataframes.items():
        json_bytes = BytesIO()
        # Ensure dates are formatted as strings
        df = df.copy()
        for col in df.columns:
            if 'Date' in col or 'FetchedAt' in col:
                df[col] = df[col].astype(str)
        df.to_json(json_bytes, orient='records', date_format='iso')
        json_bytes.seek(0)
        json_file_name = f"{base_filename}-{df_name}.json"
        azure_blob_url = upload_to_azure_blob_stream(
            connection_string,
            container_name,
            json_bytes,
            json_file_name,
            directory='json/',
            use_managed_identity=use_managed_identity
        )
        logger.info(f"JSON file uploaded to Azure Blob Storage: {azure_blob_url}")

def upload_excel_to_azure_blob(connection_string, dataframes, base_filename, container_name, use_managed_identity):
    """Upload Excel dataframes to Azure Blob Storage."""
    excel_bytes = create_excel_file(dataframes)
    excel_file_name = f"{base_filename}.xlsx"
    azure_blob_url = upload_to_azure_blob_stream(
        connection_string,
        container_name,
        excel_bytes,
        excel_file_name,
        directory='excel/',
        use_managed_identity=use_managed_identity
    )
    logger.info(f"Excel file uploaded to Azure Blob Storage: {azure_blob_url}")

def save_files_locally(dataframes, base_filename, output_format):
    """Save dataframes to local files."""
    output_directory = "output"
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    if output_format in ['excel', 'all']:
        excel_file_path = os.path.join(output_directory, f"{base_filename}.xlsx")
        excel_bytes = create_excel_file(dataframes)
        with open(excel_file_path, 'wb') as f:
            f.write(excel_bytes.getvalue())
        logger.info(f"Excel file saved locally at: {excel_file_path}")

    if output_format in ['json', 'all']:
        for df_name, df in dataframes.items():
            json_file_path = os.path.join(output_directory, f"{base_filename}-{df_name}.json")
            # Ensure dates are formatted as strings
            df = df.copy()
            for col in df.columns:
                if 'Date' in col or 'FetchedAt' in col:
                    df[col] = df[col].astype(str)
            df.to_json(json_file_path, orient='records', date_format='iso')
            logger.info(f"JSON file saved locally at: {json_file_path}")

def create_excel_file(dataframes):
    """Create an Excel file from dataframes and return it as BytesIO."""
    excel_bytes = BytesIO()
    with pd.ExcelWriter(excel_bytes, engine='openpyxl') as writer:
        for df_name, df in dataframes.items():
            # Ensure date columns are formatted as strings in MM-DD-YYYY format
            df = df.copy()
            for col in df.columns:
                if 'Date' in col or 'FetchedAt' in col:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%m-%d-%Y')
            df.to_excel(writer, sheet_name=df_name, index=False)
            format_excel_sheet(writer, df_name, df)
    excel_bytes.seek(0)
    return excel_bytes

def format_excel_sheet(writer, sheet_name, df):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    header_font = Font(bold=True)
    for cell in worksheet['1:1']:  # First row is the header
        cell.font = header_font

    # Adjust column widths
    for col_num, column_title in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        # Ensure all data is converted to string before measuring length
        df[column_title] = df[column_title].astype(str)
        column_width = max(df[column_title].map(len).max(), len(column_title))
        worksheet.column_dimensions[column_letter].width = column_width + 2

    # Format date columns without using NamedStyle
    for col in df.columns:
        if 'Date' in col or 'FetchedAt' in col:
            col_idx = df.columns.get_loc(col) + 1  # Adjust for zero-based index
            for cell in worksheet.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in cell:
                    c.number_format = 'MM-DD-YYYY'  # Set date format directly

def upload_to_azure_blob_stream(connection_string, container_name, stream, blob_name, directory='', use_managed_identity=False):
    try:
        blob_service_client = create_blob_service_client(connection_string, use_managed_identity)
        container_client = blob_service_client.get_container_client(container_name)

        try:
            container_client.create_container()
        except ResourceExistsError:
            logger.info(f"Container '{container_name}' already exists.")
        except Exception as e:
            logger.error(f"Error creating container: {e}")
            return

        full_blob_name = f"{directory}{blob_name}" if directory else blob_name
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=full_blob_name)
        blob_client.upload_blob(stream, overwrite=True)

        if not use_managed_identity:
            sas_token = generate_blob_sas(
                account_name=blob_service_client.account_name,
                container_name=container_name,
                blob_name=full_blob_name,
                account_key=blob_service_client.credential.account_key,
                permission=BlobSasPermissions(read=True),
                expiry=datetime.utcnow() + timedelta(hours=24)
            )
            return f"https://{blob_service_client.account_name}.blob.core.windows.net/{container_name}/{full_blob_name}?{sas_token}"
        else:
            return f"https://{blob_service_client.account_name}.blob.core.windows.net/{container_name}/{full_blob_name}"
    except Exception as e:
        logger.error(f"An error occurred while uploading to Azure Blob Storage: {e}")

def create_blob_service_client(connection_string, use_managed_identity=False):
    if use_managed_identity:
        credential = DefaultAzureCredential()
        return BlobServiceClient(account_url=connection_string, credential=credential)
    return BlobServiceClient.from_connection_string(connection_string)

def get_github_data(api_url, token):
    headers = {'Authorization': f'token {token}'}
    response = requests.get(api_url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        logger.error(f"Failed to fetch data: {response.status_code} - {response.text}")
        return None

def get_github_repo_info(api_url, token):
    headers = {'Authorization': f'token {token}'}
    response = requests.get(api_url, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        logger.error(f"Failed to fetch repository info: {response.status_code} - {response.text}")
        return None

def get_stars_data(api_url, token):
    stars_data = []
    page = 1
    while True:
        headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3.star+json'
        }
        url = f"{api_url}/stargazers?page={page}&per_page=100"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if data:
                stars_data.extend(data)
                page += 1
            else:
                break
        else:
            logger.error(f"Failed to fetch stars data: {response.status_code} - {response.text}")
            break
    return stars_data

def get_forks_data(api_url, token):
    forks_data = []
    page = 1
    while True:
        headers = {
            'Authorization': f'token {token}',
            'Accept': 'application/vnd.github.v3+json'
        }
        url = f"{api_url}/forks?page={page}&per_page=100"
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if data:
                forks_data.extend(data)
                page += 1
            else:
                break
        else:
            logger.error(f"Failed to fetch forks data: {response.status_code} - {response.text}")
            break
    return forks_data

def process_about_data(repo_info, owner, repo):
    about_text = repo_info.get('description', None)
    return {
        "Repo Owner": owner,
        "Repo Name": repo,
        "About": about_text
    }

def process_stars_data(stars_data, owner, repo, db_client, db_type):
    """Process stars data and return the incremental difference of stars each day."""
    cumulative_count = {}
    total_stars = 0
    sorted_data = sorted(stars_data, key=lambda x: x['starred_at'])

    for star_info in sorted_data:
        date = star_info['starred_at'].split('T')[0]
        total_stars += 1
        cumulative_count[date] = total_stars

    processed_data = [
        {"Repo Owner": owner,
         "Repo Name": repo,
         "Date": datetime.strptime(date, '%Y-%m-%d').strftime('%m-%d-%Y'),
         "Total Stars": cumulative_count[date] - (cumulative_count.get(date, 0) - 1)}  # Mapping Stars Added to Total Stars
        for date in cumulative_count
    ]
    
    return processed_data


def process_forks_data(forks_data, owner, repo, db_client, db_type):
    """Process forks data and return the incremental difference of forks each day."""
    cumulative_count = {}
    total_forks = 0
    sorted_data = sorted(forks_data, key=lambda x: x['created_at'])

    for fork_info in sorted_data:
        date = fork_info['created_at'].split('T')[0]
        total_forks += 1
        cumulative_count[date] = total_forks

    processed_data = [
        {"Repo Owner": owner,
         "Repo Name": repo,
         "Date": datetime.strptime(date, '%Y-%m-%d').strftime('%m-%d-%Y'),
         "Total Forks": cumulative_count[date] - (cumulative_count.get(date, 0) - 1)}  # Mapping Forks Added to Total Forks
        for date in cumulative_count
    ]
    
    return processed_data


def process_traffic_data(data, owner, repo):
    if not data or 'views' not in data:
        logger.warning("No traffic data available.")
        return []
    return [
        {"Repo Owner": owner,
         "Repo Name": repo,
         "Date": datetime.strptime(item.get('timestamp', '')[:10], '%Y-%m-%d').strftime('%m-%d-%Y') if item.get('timestamp') else None,
         "Views": item.get('count', None),
         "Unique visitors": item.get('uniques', None)}
        for item in data.get('views', [])
    ]

def process_clones_data(data, owner, repo):
    if not data or 'clones' not in data:
        logger.warning("No clones data available.")
        return []
    return [
        {"Repo Owner": owner,
         "Repo Name": repo,
         "Date": datetime.strptime(item.get('timestamp', '')[:10], '%Y-%m-%d').strftime('%m-%d-%Y') if item.get('timestamp') else None,
         "Clones": item.get('count', None),
         "Unique cloners": item.get('uniques', None)}
        for item in data.get('clones', [])
    ]

def process_referrers_data(data, owner, repo):
    if data is None:
        return []
    timestamp = datetime.now().strftime('%m-%d-%Y')
    return [
        {"Repo Owner": owner,
         "Repo Name": repo,
         "Referring site": item.get('referrer', None),
         "Views": item.get('count', None),
         "Unique visitors": item.get('uniques', None),
         "FetchedAt": timestamp}
        for item in data
    ]

def process_popular_content_data(data, owner, repo):
    if data is None:
        return []
    timestamp = datetime.now().strftime('%m-%d-%Y')
    return [
        {"Repo Owner": owner,
         "Repo Name": repo,
         "Path": item.get('path', None),
         "Title": item.get('title', None),
         "Views": item.get('count', None),
         "Unique visitors": item.get('uniques', None),
         "FetchedAt": timestamp}
        for item in data
    ]

def read_token_from_file(file_path):
    try:
        with open(file_path, 'r') as file:
            return file.read().strip()
    except IOError as e:
        logger.error(f"Error reading token file: {e}")
        return None

def main():
    parser = argparse.ArgumentParser(description='GitHub Repository Traffic Data Fetcher')
    parser.add_argument('--repo', required=True, help='Repository name')
    parser.add_argument('--owner', required=True, help='Organization/user name that owns the repository')
    parser.add_argument('--output-format', choices=['excel', 'json', 'all'], default='excel', help='Output format for the data (excel, json, or all)')
    parser.add_argument('--filename', help='Optional: Specify a filename for the output. If not provided, defaults to {owner}-{repo}-traffic-data')
    parser.add_argument('--token-file', required=True, help='Path to a text file containing the GitHub Personal Access Token')
    parser.add_argument('--db-connection-string', required=True, help='Database connection string to store and retrieve the data')
    parser.add_argument('--db-type', choices=['mongodb', 'cosmosdb'], required=True, help='Type of database (mongodb or cosmosdb)')
    parser.add_argument('--azure-storage-connection-string', help='Optional: Azure Blob Storage connection string for storing the output file')
    parser.add_argument('--managed-identity-storage', required=False, action='store_true', help='Use Managed Identity for Azure Blob Storage authentication')

    args = parser.parse_args()

    token = read_token_from_file(args.token_file)
    if not token:
        logger.error("Failed to read GitHub token.")
        return

    retrieve_and_process_stats(args.owner, args.repo, args.filename, args.db_connection_string, args.db_type, args.azure_storage_connection_string, args.output_format, token, args.managed_identity_storage)

if __name__ == "__main__":
    main()
